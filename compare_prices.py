import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import os
import glob

# Создаем директории, если они не существуют
os.makedirs('input/old', exist_ok=True)
os.makedirs('input/new', exist_ok=True)
os.makedirs('result', exist_ok=True)

# Находим первый файл в директории old и new
input_files_old = glob.glob('input/old/*.*')
input_files_new = glob.glob('input/new/*.*')

# Проверяем наличие входных файлов
if not input_files_old:
    print("Ошибка: В папке 'input/old' отсутствуют файлы прайс-листа предыдущего года")
    exit(1)

if not input_files_new:
    print("Ошибка: В папке 'input/new' отсутствуют файлы прайс-листа следующего года")
    exit(1)

# Выбираем первый файл в каждой директории
input_file_old = input_files_old[0]
input_file_new = input_files_new[0]

# Определяем годы для имен колонок из имен файлов
old_year = "предыдущий год"
new_year = "следующий год"

# Формируем название выходного файла
output_file = os.path.join('result', f'Сравнение прайсов ({os.path.basename(input_file_new)}).xlsx')

# Выводим информацию о файлах
print(f"Используем файлы:")
print(f"Прайс предыдущего года: {input_file_old}")
print(f"Прайс следующего года: {input_file_new}")
print(f"Результат будет сохранен в: {output_file}")

# Загрузка файлов
try:
    # Читаем Excel файлы
    engine = 'xlrd' if input_file_old.endswith('.xls') else 'openpyxl'
    df_old = pd.read_excel(
        input_file_old,
        engine=engine,  # Выбираем движок в зависимости от расширения
        na_values=['', ' ', None],  # Считаем пустые строки и пробелы как NaN
        keep_default_na=True
    )
    
    engine = 'xlrd' if input_file_new.endswith('.xls') else 'openpyxl'
    df_new = pd.read_excel(
        input_file_new,
        engine=engine,  # Выбираем движок в зависимости от расширения
        na_values=['', ' ', None],  # Считаем пустые строки и пробелы как NaN
        keep_default_na=True
    )
    
    # Выводим информацию о столбцах для отладки
    print("\nСтолбцы в старом прайсе:")
    for col in df_old.columns:
        print(f"- {col}")

    print("\nСтолбцы в новом прайсе:")
    for col in df_new.columns:
        print(f"- {col}")
    
    # Переименовываем колонки в обоих DataFrame
    column_names = {
        df_old.columns[0]: '№ услуги',
        df_old.columns[1]: 'Артикул',
        df_old.columns[2]: 'Наименование услуги',
        df_old.columns[-1]: f'Стоимость услуг {old_year}'
    }
    df_old = df_old.rename(columns=column_names)
    
    column_names = {
        df_new.columns[0]: '№ услуги',
        df_new.columns[1]: 'Артикул',
        df_new.columns[2]: 'Наименование услуги',
        df_new.columns[-1]: f'Стоимость услуг {new_year}'
    }
    df_new = df_new.rename(columns=column_names)
    
    # Обновляем имена колонок с ценами
    price_column_old = f'Стоимость услуг {old_year}'
    price_column_new = f'Стоимость услуг {new_year}'
    
    print(f"\nИспользуем колонки:")
    print(f"Предыдущий год: {price_column_old}")
    print(f"Следующий год: {price_column_new}")
    
    # Преобразуем значения цен в числовой формат
    def clean_price(value):
        if pd.isna(value):
            return value
        try:
            # Если значение уже числовое, просто округляем и возвращаем
            if isinstance(value, (int, float)):
                return round(float(value), 2)
            
            # Преобразуем в строку и очищаем
            value_str = str(value)
            # Проверяем, является ли значение числом в научной нотации
            if 'e' in value_str.lower():
                return float(value_str)
            # Удаляем все пробелы (в начале, в конце и в середине)
            value_str = ''.join(value_str.split())
            # Удаляем апостроф в начале, если он есть
            value_str = value_str.lstrip("'")
            # Заменяем запятые на точки
            value_str = value_str.replace(',', '.')
            # Пробуем преобразовать в число
            if value_str:
                num = float(value_str)
                # Округляем до 2 знаков после запятой для устранения ошибок округления
                return round(num, 2)
            return None
        except (ValueError, TypeError):
            # Выводим проблемное значение для отладки
            print(f"Не удалось преобразовать значение: '{value}' (тип: {type(value)})")
            return None

    # Применяем очистку к колонкам с ценами и выводим уникальные значения для отладки
    print("\nУникальные значения в колонке цен предыдущего года:")
    print(df_old[price_column_old].unique())
    
    print("\nУникальные значения в колонке цен следующего года:")
    print(df_new[price_column_new].unique())
    
    # Применяем очистку к колонкам с ценами
    df_old[price_column_old] = df_old[price_column_old].apply(clean_price)
    df_new[price_column_new] = df_new[price_column_new].apply(clean_price)
    
    # Проверяем наличие цен в новом прайсе
    print("\nПроверка цен в новом прайсе (после очистки):")
    for index, row in df_new.iterrows():
        article = str(row['Артикул']) if pd.notna(row['Артикул']) else ""
        if article.strip() and pd.isna(row[price_column_new]):
            name = row['Наименование услуги'] if pd.notna(row['Наименование услуги']) else "Нет названия"
            print(f"ВНИМАНИЕ: Отсутствует цена для артикула {article} ({name})")
    
    # Выводим результаты преобразования для проверки
    print("\nПосле преобразования:")
    print("Уникальные значения в колонке цен предыдущего года:")
    print(df_old[price_column_old].unique())
    
    print("\nУникальные значения в колонке цен следующего года:")
    print(df_new[price_column_new].unique())
    
    # Удаляем пустые строки
    # Строка считается пустой, если все значения в ней NaN или пустые строки
    df_old = df_old.dropna(how='all').reset_index(drop=True)
    df_new = df_new.dropna(how='all').reset_index(drop=True)
    
    # Удаляем строки, где все значения - пустые строки
    df_old = df_old[~(df_old.astype(str) == '').all(axis=1)].reset_index(drop=True)
    df_new = df_new[~(df_new.astype(str) == '').all(axis=1)].reset_index(drop=True)
    
    # НЕ фильтруем строки без артикулов, чтобы сохранить заголовки разделов
    # Заголовки разделов могут не иметь артикулов, но быть нужными для структуры прайса
    
    # Выводим статистику по количеству позиций
    print("\nСтатистика по количеству позиций:")
    print(f"Количество позиций в прайсе предыдущего года: {len(df_old)}")
    print(f"Количество позиций в прайсе следующего года: {len(df_new)}")
    
    # Подсчитываем уникальные артикулы (конвертируем в строки для корректного сравнения)
    # Исключаем заголовки разделов из сравнения артикулов
    articles_old = set(df_old[df_old['Артикул'].notna() & (df_old['Артикул'].astype(str) != '')]['Артикул'].astype(str).dropna())
    articles_new = set(df_new[df_new['Артикул'].notna() & (df_new['Артикул'].astype(str) != '')]['Артикул'].astype(str).dropna())
    
    print(f"\nКоличество уникальных артикулов:")
    print(f"Прайс предыдущего года: {len(articles_old)}")
    print(f"Прайс следующего года: {len(articles_new)}")
    
    # Анализируем разницу по артикулам
    missing_in_new = articles_old - articles_new
    new_in_new = articles_new - articles_old
    
    print(f"\nАнализ изменений по артикулам:")
    print(f"Услуг, отсутствующих в новом прайсе: {len(missing_in_new)}")
    if len(missing_in_new) > 0:
        print("\nСписок артикулов, отсутствующих в новом прайсе:")
        for article in sorted(missing_in_new):
            # Конвертируем артикул в строку для поиска
            service = df_old[df_old['Артикул'].astype(str) == article]['Наименование услуги'].iloc[0]
            price = df_old[df_old['Артикул'].astype(str) == article][price_column_old].iloc[0]
            print(f"Артикул: {article}, Цена предыдущего года: {price} руб., Услуга: {service}")
    
    print(f"\nНовых услуг в новом прайсе: {len(new_in_new)}")
    if len(new_in_new) > 0:
        print("\nСписок новых артикулов в новом прайсе:")
        for article in sorted(new_in_new):
            # Конвертируем артикул в строку для поиска
            service = df_new[df_new['Артикул'].astype(str) == article]['Наименование услуги'].iloc[0]
            price = df_new[df_new['Артикул'].astype(str) == article][price_column_new].iloc[0]
            print(f"Артикул: {article}, Цена нового прайса: {price} руб., Услуга: {service}")
    
    # Сохраняем список удаленных позиций для добавления в конец документа
    removed_services = []
    if len(missing_in_new) > 0:
        for article in sorted(missing_in_new):
            # Проверяем, что артикул не пустой и существует в старом прайсе
            if article and article.strip() and (df_old['Артикул'].astype(str) == article).any():
                try:
                    service_num = df_old[df_old['Артикул'].astype(str) == article]['№ услуги'].iloc[0]
                    service = df_old[df_old['Артикул'].astype(str) == article]['Наименование услуги'].iloc[0]
                    price = df_old[df_old['Артикул'].astype(str) == article][price_column_old].iloc[0]
                    # Добавляем только если имеем все данные
                    if pd.notna(service_num) and pd.notna(service) and pd.notna(price):
                        removed_services.append((service_num, article, price, service))
                except (IndexError, KeyError) as e:
                    print(f"Ошибка при обработке артикула {article}: {e}")
        
        # Сортируем список по номеру услуги
        # Преобразуем номер услуги в число для корректной сортировки
        removed_services.sort(key=lambda x: float(str(x[0]).replace(',', '.')) if str(x[0]).replace(',', '.').replace('.', '').isdigit() else float('inf'))

except Exception as e:
    print(f"Ошибка при загрузке файлов: {e}")
    exit(1)

# Создаем словарь цен предыдущего года для быстрого поиска по артикулу (конвертируем в строки)
# Отфильтровываем записи без артикулов и пустые артикулы
valid_old_records = df_old[df_old['Артикул'].notna() & (df_old['Артикул'].astype(str).str.strip() != '')]
prices_old = dict(zip(valid_old_records['Артикул'].astype(str), valid_old_records[price_column_old]))

# Добавляем колонку с ценами предыдущего года, используя артикул для сопоставления
df_new['Стоимость услуг предыдущий год'] = df_new['Артикул'].astype(str).map(prices_old)

# Вычисляем процентное изменение
def calculate_price_change(row):
    article = row['Артикул']
    price_new = row[price_column_new]
    price_old = prices_old.get(str(article), None)
    
    if price_old is not None and price_old != 0 and pd.notnull(price_old) and pd.notnull(price_new):
        change = ((price_new - price_old) / price_old) * 100
        return round(change, 1)
    return None

# Добавляем новую колонку с процентным изменением
df_new['Изменение цены %'] = df_new.apply(calculate_price_change, axis=1)

# Форматируем процентное изменение для отображения
df_new['Изменение цены % (текст)'] = df_new['Изменение цены %'].apply(
    lambda x: f"+{x:.1f}%" if pd.notnull(x) and x > 0 else (f"{x:.1f}%" if pd.notnull(x) else "")
)

# Переупорядочиваем колонки
columns = list(df_new.columns)
price_new_index = columns.index(price_column_new)
columns.remove('Стоимость услуг предыдущий год')
columns.insert(price_new_index, 'Стоимость услуг предыдущий год')
df_new = df_new[columns]

# ВАЖНОЕ ИСПРАВЛЕНИЕ: Копируем DataFrame вместо создания нового
df_result = df_new.copy(deep=True)

# Проверяем, что цены скопировались корректно
print("\nПроверка цен в результирующем DataFrame перед сохранением:")
for index, row in df_result.iterrows():
    article = str(row['Артикул']) if pd.notna(row['Артикул']) else ""
    if article.strip():
        # Проверяем, есть ли цена
        if pd.isna(row[price_column_new]):
            name = row['Наименование услуги'] if pd.notna(row['Наименование услуги']) else "Нет названия"
            print(f"ВНИМАНИЕ: Отсутствует цена для артикула {article} ({name})")
            
            # Попытка восстановить цену из исходного DataFrame
            orig_row = df_new[df_new['Артикул'].astype(str) == article]
            if not orig_row.empty and pd.notna(orig_row[price_column_new].iloc[0]):
                corrected_price = orig_row[price_column_new].iloc[0]
                print(f"  Восстанавливаем цену: {corrected_price}")
                df_result.at[index, price_column_new] = corrected_price

# Фильтруем результат, оставляя только:
# 1. Строки без артикулов (заголовки)
# 2. Строки с артикулами, которые есть в новом прайсе
df_result = df_result[
    (~df_result['Артикул'].notna()) |  # Строки без артикулов
    (df_result['Артикул'].notna() & df_result['Артикул'].astype(str).isin(articles_new))  # Строки с артикулами из нового прайса
]

# Сохраняем результат в новый файл Excel
df_result.to_excel(output_file, index=False)

# Добавляем цветовое форматирование
wb = load_workbook(output_file)
ws = wb.active

# Определяем цвета для форматирования
olive_fill = PatternFill(start_color='E6F2D5', end_color='E6F2D5', fill_type='solid')  # Светлый оливковый
light_blue_fill = PatternFill(start_color='DEEAF6', end_color='DEEAF6', fill_type='solid')  # Светлый синий
yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Более мягкий желтый
light_gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')  # Светло-серый
black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')  # Черный
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # Белый

# Определяем стиль границ
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Определяем белые границы для ячеек вне таблицы
white_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)

# Находим индексы нужных колонок
price_col_new = None
price_col_old = None
percent_col = None
percent_text_col = None
service_name_col = None  # Добавляем поиск колонки с наименованием услуги

for idx, cell in enumerate(ws[1], 1):
    if cell.value == price_column_new:
        price_col_new = idx
    elif cell.value == 'Стоимость услуг предыдущий год':
        price_col_old = idx
    elif cell.value == 'Изменение цены %':
        percent_col = idx
    elif cell.value == 'Изменение цены % (текст)':
        percent_text_col = idx
    elif cell.value == 'Наименование услуги':
        service_name_col = idx

# Применяем форматирование
if price_col_new and percent_col:
    # Получаем размеры таблицы
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Форматируем заголовки колонок (первая строка)
    for col in range(1, max_col + 1):
        header_cell = ws.cell(row=1, column=col)
        header_cell.border = thin_border
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(
            wrap_text=True,
            horizontal='center',
            vertical='center',
            shrink_to_fit=False,
            indent=0
        )
        # Добавляем отступы в текст заголовка
        if header_cell.value:
            header_cell.value = f" {header_cell.value} "
    
    # Устанавливаем автоматическую высоту для строки заголовков
    ws.row_dimensions[1].height = None
    
    # Применяем форматирование к остальным ячейкам
    for row in range(2, max_row + 1):  # Начинаем со второй строки
        # Устанавливаем автоматическую высоту строки
        ws.row_dimensions[row].height = None
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            # Добавляем границы для всех ячеек
            cell.border = thin_border
            
            # Настраиваем форматирование в зависимости от типа колонки
            if col == service_name_col:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical='center',
                    horizontal='left',
                    shrink_to_fit=False,
                    indent=0
                )
                # Добавляем пробелы в начало и конец текста для создания отступов
                if cell.value and isinstance(cell.value, str):
                    cell.value = f" {cell.value} "
            else:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    shrink_to_fit=False,
                    indent=0
                )
                # Форматируем ячейки с ценами
                if col in [price_col_old, price_col_new]:
                    cell.number_format = '0.00'
                # Добавляем пробелы в начало и конец текста для создания отступов (кроме цен)
                elif cell.value and isinstance(cell.value, (str, int, float)):
                    cell.value = f" {cell.value} "
            
            # Применяем цветовое форматирование для колонок с процентами
            if col == percent_col or col == percent_text_col:
                try:
                    if col == percent_col:
                        value = float(cell.value) if isinstance(cell.value, (int, float)) else None
                    else:  # percent_text_col
                        # Извлекаем число из текстового представления
                        value = float(cell.value.strip(' %+')) if cell.value and cell.value.strip(' %+-').replace('.', '').isdigit() else None
                    
                    if value is not None:
                        if value > 5:  # Значительное повышение
                            cell.fill = olive_fill
                        elif value < -5:  # Значительное снижение
                            cell.fill = light_blue_fill
                        elif value != 0:  # Небольшое изменение
                            cell.fill = yellow_fill
                except (ValueError, TypeError):
                    pass  # Пропускаем ячейки с некорректными значениями
        
        # Проверяем цену
        price_cell = ws.cell(row=row, column=price_col_new)
        
        # Если цена отсутствует
        if price_cell.value is None or price_cell.value == '':
            # Проверяем, содержит ли строка информацию только в одной ячейке
            non_empty_cells = 0
            first_non_empty_cell = None
            
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).strip():
                    non_empty_cells += 1
                    if first_non_empty_cell is None:
                        first_non_empty_cell = cell
            
            # Если информация только в одной ячейке
            if non_empty_cells == 1 and first_non_empty_cell:
                # Сохраняем значение из непустой ячейки
                cell_value = first_non_empty_cell.value.strip()  # Убираем лишние пробелы
                
                # Очищаем все ячейки в строке
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).value = None
                
                # Объединяем все ячейки в строке в одну
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
                
                # Форматируем объединенную ячейку
                merged_cell = ws.cell(row=row, column=1)
                merged_cell.value = f" {cell_value} "  # Добавляем отступы
                merged_cell.fill = black_fill
                merged_cell.font = Font(bold=True, color='FFFFFF')
                merged_cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    shrink_to_fit=False,
                    indent=0
                )
                merged_cell.border = thin_border
            else:
                # Для остальных строк без цен - только полужирный шрифт
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)

# Устанавливаем ширину колонок с учетом отступов
ws.column_dimensions[chr(64 + service_name_col)].width = 65  # Увеличиваем ширину для учета отступов
for col in range(1, max_col + 1):
    if col != service_name_col:
        column_letter = chr(64 + col)
        current_width = ws.column_dimensions[column_letter].width
        if not current_width:
            current_width = 8.43  # Стандартная ширина Excel
        ws.column_dimensions[column_letter].width = current_width + 2  # Добавляем место для отступов

# Добавляем легенду в конец документа
# Сначала добавим пустую строку
last_row = ws.max_row + 2

# Добавляем заголовок легенды
legend_header = ws.cell(row=last_row, column=1, value="Легенда цветового обозначения:")
legend_header.font = Font(bold=True)
legend_header.border = thin_border
legend_header.fill = white_fill
last_row += 2

# Добавляем описание цветов
legend_items = [
    (olive_fill, "Повышение цены более чем на 5%"),
    (light_blue_fill, "Снижение цены более чем на 5%"),
    (yellow_fill, "Изменение цены в пределах ±5%")
]

for fill, description in legend_items:
    # Создаем ячейку с цветом
    color_cell = ws.cell(row=last_row, column=1, value="")
    color_cell.fill = fill
    color_cell.border = thin_border
    
    # Добавляем описание
    desc_cell = ws.cell(row=last_row, column=2, value=description)
    desc_cell.alignment = Alignment(vertical='center')
    desc_cell.border = thin_border
    desc_cell.fill = white_fill
    
    last_row += 1

# Устанавливаем ширину колонок для легенды
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 40

# Добавляем список удаленных услуг, если они есть
if removed_services:
    last_row += 3  # Добавляем три пустые строки после легенды
    
    # Добавляем заголовок для списка удаленных позиций с ярким форматированием
    header = ws.cell(row=last_row, column=1, value="Список позиций, отсутствующих в прайсе 2025 года:")
    header.font = Font(bold=True, size=14, color="FF0000")  # Красный цвет, крупный шрифт
    header.border = thin_border
    header.fill = light_gray_fill
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)  # Увеличиваем до 4 колонок
    
    last_row += 2  # Пропускаем строку после заголовка
    
    # Добавляем заголовки колонок в новом порядке с ярким форматированием
    headers = ['№ услуги', 'Артикул', 'Наименование услуги', 'Цена предыдущего года']
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=last_row, column=col, value=header_text)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = yellow_fill  # Используем желтый цвет для лучшей видимости
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Добавляем данные в новом порядке
    for service_num, article, price, service in removed_services:
        last_row += 1
        # № услуги
        ws.cell(row=last_row, column=1, value=service_num).border = thin_border
        # Артикул
        ws.cell(row=last_row, column=2, value=article).border = thin_border
        # Наименование услуги
        service_cell = ws.cell(row=last_row, column=3, value=service)
        service_cell.border = thin_border
        service_cell.alignment = Alignment(horizontal='left', vertical='center')
        # Цена
        price_cell = ws.cell(row=last_row, column=4, value=price)
        price_cell.border = thin_border
        price_cell.number_format = '0.00'  # Формат с двумя десятичными знаками
    
    # Устанавливаем ширину колонок для списка удаленных позиций
    ws.column_dimensions['A'].width = 15  # Для номера услуги
    ws.column_dimensions['B'].width = 15  # Для артикула
    ws.column_dimensions['C'].width = 65  # Для наименования услуги
    ws.column_dimensions['D'].width = 15  # Для цены

# Добавляем список новых позиций (которые есть в новом прайсе, но отсутствуют в старом)
# Подготавливаем список новых позиций
new_services = []
if len(new_in_new) > 0:
    for article in sorted(new_in_new):
        # Проверяем, что артикул не пустой и существует в новом прайсе
        if article and article.strip() and (df_new['Артикул'].astype(str) == article).any():
            try:
                service_num = df_new[df_new['Артикул'].astype(str) == article]['№ услуги'].iloc[0]
                service = df_new[df_new['Артикул'].astype(str) == article]['Наименование услуги'].iloc[0]
                price = df_new[df_new['Артикул'].astype(str) == article][price_column_new].iloc[0]
                # Добавляем только если имеем все данные
                if pd.notna(service_num) and pd.notna(service) and pd.notna(price):
                    new_services.append((service_num, article, price, service))
            except (IndexError, KeyError) as e:
                print(f"Ошибка при обработке нового артикула {article}: {e}")
    
    # Сортируем список по номеру услуги
    # Преобразуем номер услуги в число для корректной сортировки
    new_services.sort(key=lambda x: float(str(x[0]).replace(',', '.')) if str(x[0]).replace(',', '.').replace('.', '').isdigit() else float('inf'))

# Если есть новые позиции, добавляем их в результат
if new_services:
    last_row += 3  # Добавляем три пустые строки перед новым разделом
    
    # Добавляем заголовок для списка новых позиций с ярким форматированием
    header = ws.cell(row=last_row, column=1, value="СПИСОК НОВЫХ ПОЗИЦИЙ, ПОЯВИВШИХСЯ В НОВОМ ПРАЙСЕ:")
    header.font = Font(bold=True, size=14, color="008000")  # Зеленый цвет, крупный шрифт
    header.border = thin_border
    header.fill = light_gray_fill
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)  # Увеличиваем до 4 колонок
    
    last_row += 2  # Пропускаем строку после заголовка
    
    # Добавляем заголовки колонок в новом порядке с ярким форматированием
    headers = ['№ услуги', 'Артикул', 'Наименование услуги', 'Цена нового прайса']
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=last_row, column=col, value=header_text)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = olive_fill  # Используем оливковый цвет для лучшей видимости
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Добавляем данные в новом порядке
    for service_num, article, price, service in new_services:
        last_row += 1
        # № услуги
        ws.cell(row=last_row, column=1, value=service_num).border = thin_border
        # Артикул
        ws.cell(row=last_row, column=2, value=article).border = thin_border
        # Наименование услуги
        service_cell = ws.cell(row=last_row, column=3, value=service)
        service_cell.border = thin_border
        service_cell.alignment = Alignment(horizontal='left', vertical='center')
        # Цена
        price_cell = ws.cell(row=last_row, column=4, value=price)
        price_cell.border = thin_border
        price_cell.number_format = '0.00'  # Формат с двумя десятичными знаками
    
    # Сохраняем те же настройки ширины колонок, что и для удаленных позиций

# Вычисляем общую разницу цен (только для позиций, которые есть в обоих прайсах)
total_old = 0
total_new = 0

# Используем только те позиции, которые есть в обоих прайсах (исключаем строки с NaN в ценах)
df_common = df_new[
    (df_new['Стоимость услуг предыдущий год'].notna()) & 
    (df_new[price_column_new].notna()) &
    (df_new['Стоимость услуг предыдущий год'] > 0) &
    (df_new[price_column_new] > 0)
]

# Считаем общие суммы для позиций, которые есть в обоих прайсах
total_old_common = df_common['Стоимость услуг предыдущий год'].sum()
total_new_common = df_common[price_column_new].sum()

# Вычисляем процент изменения для общих позиций
total_change_percent = ((total_new_common - total_old_common) / total_old_common * 100) if total_old_common > 0 else 0

# Выводим информацию о количестве позиций для проверки
print(f"\nКоличество позиций для расчета изменения по общим позициям: {len(df_common)}")
print(f"Количество позиций в старом прайсе с ценами: {len(df_old[price_column_old].dropna())}")
print(f"Количество позиций в новом прайсе с ценами: {len(df_new[price_column_new].dropna())}")

# Добавляем информацию об общем изменении цен в конец документа
last_row += 3  # Добавляем три пустые строки перед итогами

header = ws.cell(row=last_row, column=1, value="Общее изменение цен (только по общим позициям):")
header.font = Font(bold=True)
header.border = thin_border
header.fill = white_fill
ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=2)

# Добавляем значение изменения
change_value = ws.cell(row=last_row, column=3, value=f"{'+' if total_change_percent > 0 else ''}{total_change_percent:.1f}%")
change_value.border = thin_border
change_value.alignment = Alignment(horizontal='left', vertical='center')

# Применяем цветовое форматирование
if total_change_percent > 5:
    change_value.fill = olive_fill
elif total_change_percent < -5:
    change_value.fill = light_blue_fill
elif total_change_percent != 0:
    change_value.fill = yellow_fill
else:
    change_value.fill = white_fill

# Добавляем информацию об общих позициях
ws.cell(row=last_row + 1, column=1, value=f"Сумма цен предыдущего года (только общие позиции): {total_old_common:,.2f} руб.").border = thin_border
ws.cell(row=last_row + 2, column=1, value=f"Сумма цен следующего года (только общие позиции): {total_new_common:,.2f} руб.").border = thin_border

# Сохраняем файл с форматированием
wb.save(output_file)

# Выводим статистику в консоль
print(f"\nОбщее изменение цен (только по общим позициям): {'+' if total_change_percent > 0 else ''}{total_change_percent:.1f}%")
print(f"Сумма цен предыдущего года (только общие позиции): {total_old_common:,.2f} руб.")
print(f"Сумма цен следующего года (только общие позиции): {total_new_common:,.2f} руб.")

# Дополнительная статистика по новым и удаленным позициям
print(f"\nСтатистика по изменениям в прайсе:")
print(f"Количество удаленных позиций: {len(removed_services)}")
print(f"Количество новых позиций: {len(new_services)}")

# Если есть новые позиции, выводим их общую стоимость
if new_services:
    total_new_services_cost = sum(price for _, _, price, _ in new_services)
    print(f"Общая стоимость новых позиций: {total_new_services_cost:,.2f} руб.")

print(f"\nАнализ завершен. Результат сохранен в файл '{output_file}'") 